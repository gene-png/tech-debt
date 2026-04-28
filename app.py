import csv
import difflib
import io
import os
import uuid

from flask import Flask, render_template, request, redirect, url_for, abort, Response, send_from_directory
from datetime import datetime
from werkzeug.utils import secure_filename

import ai_service
import storage

app = Flask(__name__)

UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "uploads")
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
ALLOWED_EXTENSIONS = {
    ".pdf", ".csv", ".tsv", ".xlsx", ".xls", ".xlsm", ".docx", ".doc",
    ".txt", ".md", ".png", ".jpg", ".jpeg", ".gif", ".json", ".xml",
    ".zip", ".pptx", ".ppt", ".vsdx", ".vsd",
}
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_BYTES

DATA_REQUEST_TYPES = [
    ("current_inventory", "Current software inventory"),
    ("license_agreements", "License agreements"),
    ("purchase_orders", "Purchase orders"),
    ("invoices", "Invoices"),
    ("renewal_notices", "Renewal notices"),
    ("saas_exports", "SaaS subscription exports"),
    ("cloud_marketplace", "Cloud marketplace purchases"),
    ("vendor_contracts", "Vendor contracts"),
    ("enterprise_license_agreements", "Enterprise license agreements"),
    ("asset_management", "Asset management exports"),
    ("endpoint_software", "Endpoint software reports"),
    ("idp_apps", "Identity provider application lists"),
    ("finance_procurement", "Finance or procurement reports"),
    ("helpdesk_catalog", "Help desk application catalog"),
    ("cyber_tool_list", "Cybersecurity tool list"),
    ("architecture", "Architecture diagrams"),
    ("network_endpoint_exports", "Network or endpoint management exports"),
    ("critical_apps", "Business critical applications list"),
]

DATA_REQUEST_STATUSES = [
    ("requested", "Requested"),
    ("received", "Received"),
    ("not_applicable", "N/A"),
    ("waived", "Waived"),
]

# Phase 5 — Identify Overlap --------------------------------------------------

DISPOSITIONS = [
    ("keep", "Keep"),
    ("consolidate", "Consolidate"),
    ("retire", "Retire"),
    ("replace", "Replace"),
    ("renegotiate", "Renegotiate"),
    ("further_review", "Further review required"),
]

REMOVAL_RISKS = [
    ("low", "Low"),
    ("medium", "Medium"),
    ("high", "High"),
    ("unknown", "Unknown"),
]

# Dispositions that imply a product would no longer cost money once acted on.
SAVINGS_DISPOSITIONS = {"retire", "replace", "consolidate"}

# Phase 7 — Technical Debt ----------------------------------------------------

# Each entry: (key, label, helper-text shown in tooltip / register).
TECH_DEBT_FLAGS = [
    ("unsupported", "Unsupported software",
     "Product is past vendor end-of-life or no longer receiving security patches."),
    ("outdated_version", "Outdated version",
     "Running a release significantly behind current; upgrade is overdue."),
    ("unused", "Unused product",
     "Licenses are paid for but very few or no active users."),
    ("duplicative", "Duplicative tool",
     "Functionally overlapping with another tool already in scope."),
    ("no_owner", "No clear owner",
     "Missing business, technical, or contract owner."),
    ("unclear_mission", "Unclear mission need",
     "The business need this product serves is undocumented or unverified."),
    ("outside_governance", "Purchased outside governance",
     "Acquired via expense card, shadow IT, or a non-IT department without review."),
    ("no_integration", "Does not integrate",
     "Lacks integration with current systems; creates manual hand-offs."),
    ("weak_security", "Weak security controls",
     "Insufficient SSO, MFA, encryption, or access controls; flagged by security team."),
    ("manual_burden", "Excessive manual work",
     "Operating the product requires meaningful manual effort that should be automated."),
    ("data_silo", "Creates a data silo",
     "Holds important data not accessible to other systems or stakeholders."),
    ("poor_adoption", "Poor adoption",
     "Targeted users have not adopted the product despite training or rollout."),
    ("high_cost_low_value", "High cost, low value",
     "Annual cost is large relative to the business value the product delivers."),
    ("renewal_no_justification", "Nearing renewal without justification",
     "Renewal is approaching and there is no documented case to renew."),
]

DEBT_SEVERITIES = [
    ("low", "Low"),
    ("medium", "Medium"),
    ("high", "High"),
]

# Phase 8 — Estimate Cost Savings ---------------------------------------------

SAVINGS_DISPOSITION_CHOICES = [
    ("retire", "Retire"),
    ("replace", "Replace"),
    ("consolidate", "Consolidate"),
    ("renegotiate", "Renegotiate"),
    ("reduce_licenses", "Reduce licenses"),
    ("other", "Other"),
]

SAVINGS_STATUSES = [
    ("proposed", "Proposed"),
    ("approved", "Approved"),
    ("rejected", "Rejected"),
]

# Phase 9 — Stakeholder Validation --------------------------------------------

STAKEHOLDER_ROLES = [
    ("it_leadership", "IT leadership"),
    ("cybersecurity", "Cybersecurity team"),
    ("finance", "Finance"),
    ("procurement", "Procurement"),
    ("business_unit_owner", "Business unit owner"),
    ("system_administrator", "System administrator"),
    ("power_user", "Power user"),
    ("compliance_legal", "Compliance / legal"),
]

STAKEHOLDER_STATUSES = [
    ("consulted", "Consulted"),
    ("agreed", "Agreed"),
    ("pushback", "Pushback"),
    ("blocked", "Blocked"),
]

VALIDATION_QUESTIONS = [
    ("who_uses", "Who uses this tool?"),
    ("business_process", "What business process depends on it?"),
    ("what_breaks", "What would break if it was removed?"),
    ("better_tool", "Is there a better enterprise tool already available?"),
    ("required_by", "Is the tool required by contract, compliance, or customer need?"),
    ("cost_justified", "Is the cost justified by the value?"),
    ("absorbed_by", "Can the function be absorbed by another product?"),
]

VALIDATION_OVERALL_STATUSES = [
    ("not_started", "Not started"),
    ("pending", "In progress"),
    ("validated", "Validated"),
    ("blocked", "Blocked"),
]

# Phase 3 — Build Inventory ---------------------------------------------------

PRODUCT_CATEGORIES = [
    "Email and collaboration",
    "Endpoint security",
    "Identity and access management",
    "Project management",
    "Customer relationship management",
    "Data analytics",
    "Cloud services",
    "Backup and recovery",
    "Vulnerability management",
    "Network monitoring",
    "Ticketing and service desk",
    "Document management",
    "Development tools",
    "AI tools",
    "Compliance and governance tools",
    "HR and finance",
    "Other",
]

DEPLOYMENT_MODELS = [
    "SaaS",
    "Cloud-hosted",
    "Server / On-premises",
    "Desktop",
    "Mobile",
    "Hybrid",
]

DATA_SENSITIVITY_LEVELS = [
    "Public",
    "Internal",
    "Confidential",
    "Restricted",
]

# (storage_key, csv/form label, type)
PRODUCT_FIELDS = [
    ("product_name", "Product name", "text"),
    ("vendor", "Vendor", "text"),
    ("version", "Version", "text"),
    ("category", "Category", "select"),
    ("business_owner", "Business owner", "text"),
    ("technical_owner", "Technical owner", "text"),
    ("contract_owner", "Contract owner", "text"),
    ("licenses_purchased", "Licenses purchased", "int"),
    ("licenses_assigned", "Licenses assigned", "int"),
    ("active_users", "Active users", "int"),
    ("cost_per_license", "Cost per license", "money"),
    ("total_annual_cost", "Total annual cost", "money"),
    ("renewal_date", "Renewal date", "date"),
    ("contract_term", "Contract term", "text"),
    ("purchase_source", "Purchase source", "text"),
    ("deployment_model", "Deployment model", "select_deploy"),
    ("primary_use_case", "Primary use case", "text"),
    ("systems_supported", "Systems supported", "text"),
    ("data_sensitivity", "Data sensitivity", "select_sens"),
    ("security_compliance", "Security or compliance relevance", "text"),
    ("known_risks", "Known risks", "textarea"),
    ("notes", "Notes", "textarea"),
]

# Aliases for fuzzy CSV header matching (lowercase, normalized).
PRODUCT_FIELD_ALIASES = {
    "product_name": ["product", "product_name", "name", "software", "application", "app", "tool"],
    "vendor": ["vendor", "publisher", "supplier", "manufacturer"],
    "version": ["version", "release"],
    "category": ["category", "type", "function"],
    "business_owner": ["business_owner", "business owner", "business contact"],
    "technical_owner": ["technical_owner", "technical owner", "it owner", "admin"],
    "contract_owner": ["contract_owner", "contract owner", "procurement contact"],
    "licenses_purchased": ["licenses_purchased", "licenses purchased", "purchased licenses", "total licenses", "licenses", "seats"],
    "licenses_assigned": ["licenses_assigned", "licenses assigned", "assigned licenses", "assigned"],
    "active_users": ["active_users", "active users", "users", "monthly active users", "mau"],
    "cost_per_license": ["cost_per_license", "cost per license", "unit cost", "price per license", "price"],
    "total_annual_cost": ["total_annual_cost", "annual cost", "total cost", "yearly cost", "spend", "annual spend"],
    "renewal_date": ["renewal_date", "renewal date", "renewal", "next renewal", "expires"],
    "contract_term": ["contract_term", "contract term", "term"],
    "purchase_source": ["purchase_source", "purchase source", "source", "purchased via", "purchased through"],
    "deployment_model": ["deployment_model", "deployment", "deployment type", "hosting"],
    "primary_use_case": ["primary_use_case", "use case", "purpose", "primary use"],
    "systems_supported": ["systems_supported", "systems supported", "systems", "integrates with"],
    "data_sensitivity": ["data_sensitivity", "sensitivity", "data classification", "classification"],
    "security_compliance": ["security_compliance", "compliance", "security relevance", "compliance relevance"],
    "known_risks": ["known_risks", "risks", "known issues"],
    "notes": ["notes", "comments", "remarks"],
}

TOOL_CATEGORIES = [
    ("cloud", "Cloud services"),
    ("saas", "SaaS applications"),
    ("desktop", "Desktop software"),
    ("server", "Server software"),
    ("security", "Security tools"),
    ("infrastructure", "Infrastructure"),
    ("network", "Network tools"),
    ("data", "Data and analytics"),
    ("dev", "Development tools"),
    ("ai", "AI tools"),
]

PHASES = [
    ("scope", "1. Define the Scope"),
    ("data_request", "2. Request Customer Data"),
    ("inventory", "3. Build Inventory"),
    ("normalize", "4. Normalize Data"),
    ("overlap", "5. Identify Overlap"),
    ("ai_review", "6. AI Assisted Comparison"),
    ("tech_debt", "7. Identify Technical Debt"),
    ("savings", "8. Estimate Cost Savings"),
    ("validation", "9. Stakeholder Validation"),
    ("recommendations", "10. Recommendations"),
    ("exec_summary", "11. Executive Summary"),
]


def _parse_lines(text):
    if not text:
        return []
    return [line.strip() for line in text.splitlines() if line.strip()]


def _parse_owners(text):
    """Parse 'Name | Role | Email' lines into structured records."""
    out = []
    for line in _parse_lines(text):
        parts = [p.strip() for p in line.split("|")]
        record = {"name": parts[0] if len(parts) > 0 else ""}
        if len(parts) > 1:
            record["role"] = parts[1]
        if len(parts) > 2:
            record["email"] = parts[2]
        out.append(record)
    return out


def _owners_to_text(owners):
    lines = []
    for o in owners or []:
        parts = [o.get("name", "")]
        if o.get("role"):
            parts.append(o["role"])
        if o.get("email"):
            parts.append(o["email"])
        lines.append(" | ".join(parts))
    return "\n".join(lines)


@app.route("/")
def home():
    engagements = storage.list_engagements()
    return render_template("home.html", engagements=engagements, phases=PHASES)


@app.route("/engagements/new", methods=["GET", "POST"])
def engagement_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        client = request.form.get("client", "").strip()
        lead = request.form.get("lead", "").strip()
        if not name or not client:
            return render_template(
                "engagement_new.html",
                error="Engagement name and client are required.",
                form=request.form,
            )
        eng = storage.new_engagement(name=name, client=client, lead=lead)
        return redirect(url_for("engagement_view", engagement_id=eng["id"]))
    return render_template("engagement_new.html", error=None, form={})


@app.route("/engagements/<engagement_id>")
def engagement_view(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    return render_template(
        "engagement_view.html",
        eng=eng,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/scope", methods=["GET", "POST"])
def engagement_scope(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)

    if request.method == "POST":
        scope = eng.get("scope", {})
        scope["business_units"] = _parse_lines(request.form.get("business_units", ""))
        scope["tool_categories"] = request.form.getlist("tool_categories")
        scope["include_corporate_card"] = bool(request.form.get("include_corporate_card"))
        scope["include_enterprise_agreements"] = bool(request.form.get("include_enterprise_agreements"))
        scope["include_department_purchases"] = bool(request.form.get("include_department_purchases"))
        scope["include_shadow_it"] = bool(request.form.get("include_shadow_it"))
        scope["renewal_window"] = request.form.get("renewal_window", "").strip()
        scope["contract_owners"] = _parse_owners(request.form.get("contract_owners", ""))
        scope["technical_owners"] = _parse_owners(request.form.get("technical_owners", ""))
        scope["business_owners"] = _parse_owners(request.form.get("business_owners", ""))
        scope["objectives"] = request.form.get("objectives", "").strip()
        scope["out_of_scope"] = request.form.get("out_of_scope", "").strip()
        scope["constraints"] = request.form.get("constraints", "").strip()
        scope["scope_notes"] = request.form.get("scope_notes", "").strip()

        action = request.form.get("action", "save")
        if action == "finalize":
            scope["finalized"] = True
            scope["finalized_at"] = datetime.utcnow().isoformat() + "Z"
            eng["phase_progress"]["scope"] = "complete"
            eng["phase_progress"]["data_request"] = "in_progress"
            eng["status"] = "data_request"
        elif action == "reopen":
            scope["finalized"] = False
            scope["finalized_at"] = None
            eng["phase_progress"]["scope"] = "in_progress"
            eng["status"] = "scope"
        else:
            if not scope.get("finalized"):
                eng["phase_progress"]["scope"] = "in_progress"

        eng["scope"] = scope
        storage.save_engagement(eng)
        if action == "finalize":
            return redirect(url_for("scope_statement", engagement_id=engagement_id))
        return redirect(url_for("engagement_scope", engagement_id=engagement_id))

    return render_template(
        "engagement_scope.html",
        eng=eng,
        tool_categories=TOOL_CATEGORIES,
        owners_to_text=_owners_to_text,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/scope/statement")
def scope_statement(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    cat_lookup = dict(TOOL_CATEGORIES)
    return render_template(
        "scope_statement.html",
        eng=eng,
        cat_lookup=cat_lookup,
    )


@app.route("/engagements/<engagement_id>/scope/statement.txt")
def scope_statement_txt(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    s = eng.get("scope", {})
    cat_lookup = dict(TOOL_CATEGORIES)

    def yn(v):
        return "Yes" if v else "No"

    lines = []
    lines.append("SOFTWARE REVIEW SCOPE STATEMENT")
    lines.append("=" * 50)
    lines.append(f"Engagement: {eng.get('name', '')}")
    lines.append(f"Client: {eng.get('client', '')}")
    lines.append(f"Lead consultant: {eng.get('lead', '')}")
    lines.append(f"Created: {eng.get('created_at', '')}")
    if s.get("finalized_at"):
        lines.append(f"Scope finalized: {s.get('finalized_at')}")
    lines.append("")
    lines.append("OBJECTIVES")
    lines.append("-" * 50)
    lines.append(s.get("objectives") or "(not set)")
    lines.append("")
    lines.append("BUSINESS UNITS IN SCOPE")
    lines.append("-" * 50)
    for bu in s.get("business_units") or []:
        lines.append(f"- {bu}")
    if not s.get("business_units"):
        lines.append("(none listed)")
    lines.append("")
    lines.append("TOOL CATEGORIES IN SCOPE")
    lines.append("-" * 50)
    for k in s.get("tool_categories") or []:
        lines.append(f"- {cat_lookup.get(k, k)}")
    if not s.get("tool_categories"):
        lines.append("(none selected)")
    lines.append("")
    lines.append("PURCHASE SOURCES IN SCOPE")
    lines.append("-" * 50)
    lines.append(f"Corporate credit card purchases: {yn(s.get('include_corporate_card'))}")
    lines.append(f"Enterprise agreements: {yn(s.get('include_enterprise_agreements'))}")
    lines.append(f"Department-level purchases: {yn(s.get('include_department_purchases'))}")
    lines.append(f"Shadow IT / non-IT purchases: {yn(s.get('include_shadow_it'))}")
    lines.append("")
    lines.append("RENEWAL WINDOW")
    lines.append("-" * 50)
    lines.append(s.get("renewal_window") or "(not set)")
    lines.append("")

    for label, key in [
        ("CONTRACT OWNERS", "contract_owners"),
        ("TECHNICAL OWNERS", "technical_owners"),
        ("BUSINESS OWNERS", "business_owners"),
    ]:
        lines.append(label)
        lines.append("-" * 50)
        people = s.get(key) or []
        if not people:
            lines.append("(none listed)")
        for p in people:
            bits = [p.get("name", "")]
            if p.get("role"):
                bits.append(p["role"])
            if p.get("email"):
                bits.append(p["email"])
            lines.append("- " + " | ".join(bits))
        lines.append("")

    lines.append("OUT OF SCOPE")
    lines.append("-" * 50)
    lines.append(s.get("out_of_scope") or "(not set)")
    lines.append("")
    lines.append("CONSTRAINTS AND ASSUMPTIONS")
    lines.append("-" * 50)
    lines.append(s.get("constraints") or "(not set)")
    lines.append("")
    lines.append("ADDITIONAL NOTES")
    lines.append("-" * 50)
    lines.append(s.get("scope_notes") or "(none)")

    body = "\n".join(lines)
    fn = f"scope-statement-{eng.get('id')}.txt"
    return Response(
        body,
        mimetype="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 2 — Request Customer Data
# ---------------------------------------------------------------------------

def _ensure_data_request(eng):
    """Initialize the Phase 2 data structure on the engagement if missing."""
    dr = eng.get("data_request")
    if dr and dr.get("documents"):
        return False
    documents = []
    for key, label in DATA_REQUEST_TYPES:
        documents.append({
            "id": uuid.uuid4().hex[:8],
            "doc_type_key": key,
            "doc_type_label": label,
            "status": "requested",
            "notes": "",
            "uploaded_files": [],
        })
    eng["data_request"] = {
        "documents": documents,
        "customer_message": "",
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _engagement_upload_dir(engagement_id, item_id):
    p = os.path.join(UPLOADS_DIR, engagement_id, item_id)
    os.makedirs(p, exist_ok=True)
    return p


def _data_request_summary(dr):
    summary = {"requested": 0, "received": 0, "not_applicable": 0, "waived": 0}
    for d in dr["documents"]:
        summary[d["status"]] = summary.get(d["status"], 0) + 1
    summary["total"] = len(dr["documents"])
    summary["files_total"] = sum(len(d.get("uploaded_files", [])) for d in dr["documents"])
    return summary


@app.route("/engagements/<engagement_id>/data-request", methods=["GET", "POST"])
def engagement_data_request(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)

    initialized = _ensure_data_request(eng)
    if initialized:
        if (eng["phase_progress"].get("scope") == "complete"
                and eng["phase_progress"].get("data_request") == "not_started"):
            eng["phase_progress"]["data_request"] = "in_progress"
        storage.save_engagement(eng)

    if request.method == "POST":
        action = request.form.get("action")
        dr = eng["data_request"]
        if action == "save_message":
            dr["customer_message"] = request.form.get("customer_message", "").strip()
        elif action == "update_item":
            item_id = request.form.get("item_id")
            for d in dr["documents"]:
                if d["id"] == item_id:
                    new_status = request.form.get("status", d["status"])
                    if new_status in dict(DATA_REQUEST_STATUSES):
                        d["status"] = new_status
                    d["notes"] = request.form.get("notes", "").strip()
                    break
        elif action == "finalize":
            dr["finalized"] = True
            dr["finalized_at"] = datetime.utcnow().isoformat() + "Z"
            eng["phase_progress"]["data_request"] = "complete"
            eng["status"] = "inventory"
            if eng["phase_progress"].get("inventory") == "not_started":
                eng["phase_progress"]["inventory"] = "in_progress"
        elif action == "reopen":
            dr["finalized"] = False
            dr["finalized_at"] = None
            eng["phase_progress"]["data_request"] = "in_progress"
            eng["status"] = "data_request"
        else:
            return "Unknown action", 400

        if not dr["finalized"] and eng["phase_progress"]["data_request"] == "not_started":
            eng["phase_progress"]["data_request"] = "in_progress"
        storage.save_engagement(eng)
        return redirect(url_for("engagement_data_request", engagement_id=engagement_id))

    summary = _data_request_summary(eng["data_request"])
    return render_template(
        "data_request.html",
        eng=eng,
        statuses=DATA_REQUEST_STATUSES,
        summary=summary,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/data-request/items/<item_id>/upload", methods=["POST"])
def engagement_data_request_upload(engagement_id, item_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_data_request(eng)
    item = next((d for d in eng["data_request"]["documents"] if d["id"] == item_id), None)
    if not item:
        abort(404)

    files = request.files.getlist("file")
    saved = 0
    for f in files:
        if not f or not f.filename:
            continue
        original = f.filename
        ext = os.path.splitext(original)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        safe = secure_filename(original) or "upload"
        file_id = uuid.uuid4().hex[:8]
        stored_filename = f"{file_id}_{safe}"
        upload_dir = _engagement_upload_dir(engagement_id, item_id)
        full_path = os.path.join(upload_dir, stored_filename)
        f.save(full_path)
        size = os.path.getsize(full_path)
        item["uploaded_files"].append({
            "id": file_id,
            "original_filename": original,
            "stored_filename": stored_filename,
            "size_bytes": size,
            "uploaded_at": datetime.utcnow().isoformat() + "Z",
            "content_type": f.mimetype,
        })
        if item["status"] == "requested":
            item["status"] = "received"
        saved += 1

    if eng["phase_progress"].get("data_request") == "not_started":
        eng["phase_progress"]["data_request"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_data_request", engagement_id=engagement_id) + f"#item-{item_id}")


@app.route("/engagements/<engagement_id>/data-request/items/<item_id>/files/<file_id>")
def engagement_data_request_download(engagement_id, item_id, file_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    item = next((d for d in eng.get("data_request", {}).get("documents", []) if d["id"] == item_id), None)
    if not item:
        abort(404)
    fr = next((fl for fl in item["uploaded_files"] if fl["id"] == file_id), None)
    if not fr:
        abort(404)
    upload_dir = os.path.join(UPLOADS_DIR, engagement_id, item_id)
    return send_from_directory(
        upload_dir,
        fr["stored_filename"],
        as_attachment=True,
        download_name=fr["original_filename"],
    )


@app.route("/engagements/<engagement_id>/data-request/items/<item_id>/files/<file_id>/delete", methods=["POST"])
def engagement_data_request_delete_file(engagement_id, item_id, file_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    item = next((d for d in eng.get("data_request", {}).get("documents", []) if d["id"] == item_id), None)
    if not item:
        abort(404)
    fr = next((fl for fl in item["uploaded_files"] if fl["id"] == file_id), None)
    if not fr:
        abort(404)
    full_path = os.path.join(UPLOADS_DIR, engagement_id, item_id, fr["stored_filename"])
    try:
        if os.path.exists(full_path):
            os.remove(full_path)
    except OSError:
        pass
    item["uploaded_files"] = [fl for fl in item["uploaded_files"] if fl["id"] != file_id]
    if not item["uploaded_files"] and item["status"] == "received":
        item["status"] = "requested"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_data_request", engagement_id=engagement_id) + f"#item-{item_id}")


@app.route("/engagements/<engagement_id>/data-request/checklist")
def engagement_data_request_checklist(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_data_request(eng):
        storage.save_engagement(eng)
    return render_template("data_request_checklist.html", eng=eng)


@app.route("/engagements/<engagement_id>/data-request/checklist.txt")
def engagement_data_request_checklist_txt(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_data_request(eng):
        storage.save_engagement(eng)
    dr = eng["data_request"]
    status_label = dict(DATA_REQUEST_STATUSES)
    lines = []
    lines.append("CUSTOMER DATA REQUEST CHECKLIST")
    lines.append("=" * 50)
    lines.append(f"Engagement: {eng.get('name')}")
    lines.append(f"Client: {eng.get('client')}")
    if eng.get("lead"):
        lines.append(f"Lead consultant: {eng['lead']}")
    lines.append("")
    if dr.get("customer_message"):
        lines.append("MESSAGE FROM CONSULTANT")
        lines.append("-" * 50)
        lines.append(dr["customer_message"])
        lines.append("")
    lines.append("REQUESTED DOCUMENTS")
    lines.append("-" * 50)
    for i, d in enumerate(dr["documents"], 1):
        sl = status_label.get(d["status"], d["status"])
        lines.append(f"{i:2d}. [ {sl:>9s} ]  {d['doc_type_label']}")
        if d.get("notes"):
            lines.append(f"      Notes: {d['notes']}")
    body = "\n".join(lines)
    fn = f"data-request-checklist-{eng.get('id')}.txt"
    return Response(
        body,
        mimetype="text/plain",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 3 — Build the Software Inventory
# ---------------------------------------------------------------------------

def _ensure_inventory(eng):
    if "inventory" in eng and "products" in eng["inventory"]:
        return False
    eng["inventory"] = {
        "products": [],
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _new_product():
    now = datetime.utcnow().isoformat() + "Z"
    p = {"id": uuid.uuid4().hex[:8], "created_at": now, "updated_at": now}
    for key, _label, _t in PRODUCT_FIELDS:
        p[key] = ""
    return p


def _coerce_int(v):
    if v is None or v == "":
        return ""
    s = str(v).replace(",", "").strip()
    if not s:
        return ""
    try:
        return int(float(s))
    except ValueError:
        return ""


def _coerce_money(v):
    if v is None or v == "":
        return ""
    s = str(v).replace("$", "").replace(",", "").strip()
    if not s:
        return ""
    try:
        return round(float(s), 2)
    except ValueError:
        return ""


def _coerce_date(v):
    """Best-effort normalize to YYYY-MM-DD; return raw if unparseable."""
    if not v:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # keep as-is if unrecognized


def _apply_form(product, form):
    """Update a product dict from a Flask form."""
    for key, _label, ftype in PRODUCT_FIELDS:
        raw = form.get(key, "").strip() if hasattr(form, "get") else (form.get(key) or "")
        if isinstance(raw, str):
            raw = raw.strip()
        if ftype == "int":
            product[key] = _coerce_int(raw)
        elif ftype == "money":
            product[key] = _coerce_money(raw)
        elif ftype == "date":
            product[key] = _coerce_date(raw)
        else:
            product[key] = raw
    # auto-fill total_annual_cost if blank but per-license + purchased present
    if product.get("total_annual_cost") in ("", None):
        try:
            cpl = float(product.get("cost_per_license") or 0)
            qty = float(product.get("licenses_purchased") or 0)
            if cpl and qty:
                product["total_annual_cost"] = round(cpl * qty, 2)
        except (TypeError, ValueError):
            pass
    product["updated_at"] = datetime.utcnow().isoformat() + "Z"
    return product


def _normalize_header(h):
    return (h or "").strip().lower().replace("-", "_").replace("  ", " ")


def _csv_header_to_field(h):
    norm = _normalize_header(h)
    for field_key, aliases in PRODUCT_FIELD_ALIASES.items():
        if norm in [_normalize_header(a) for a in aliases]:
            return field_key
    return None


def _import_csv_text(eng, text):
    """Parse CSV text and append rows as new products. Returns (added, skipped, mapping)."""
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return 0, 0, {}
    headers = rows[0]
    mapping = {}
    for idx, h in enumerate(headers):
        field = _csv_header_to_field(h)
        if field:
            mapping[idx] = field
    if not any(v == "product_name" for v in mapping.values()):
        return 0, len(rows) - 1, {}
    added = 0
    skipped = 0
    for row in rows[1:]:
        if not row:
            continue
        product = _new_product()
        for idx, value in enumerate(row):
            field = mapping.get(idx)
            if not field:
                continue
            value = (value or "").strip()
            if field in ("licenses_purchased", "licenses_assigned", "active_users"):
                product[field] = _coerce_int(value)
            elif field in ("cost_per_license", "total_annual_cost"):
                product[field] = _coerce_money(value)
            elif field == "renewal_date":
                product[field] = _coerce_date(value)
            else:
                product[field] = value
        if not product.get("product_name"):
            skipped += 1
            continue
        # auto-fill total_annual_cost
        if not product.get("total_annual_cost"):
            try:
                cpl = float(product.get("cost_per_license") or 0)
                qty = float(product.get("licenses_purchased") or 0)
                if cpl and qty:
                    product["total_annual_cost"] = round(cpl * qty, 2)
            except (TypeError, ValueError):
                pass
        eng["inventory"]["products"].append(product)
        added += 1
    return added, skipped, mapping


def _inventory_summary(inv):
    products = inv.get("products", [])
    total = len(products)
    annual = 0.0
    licenses = 0
    assigned = 0
    by_category = {}
    for p in products:
        try:
            annual += float(p.get("total_annual_cost") or 0)
        except (TypeError, ValueError):
            pass
        try:
            licenses += int(p.get("licenses_purchased") or 0)
        except (TypeError, ValueError):
            pass
        try:
            assigned += int(p.get("licenses_assigned") or 0)
        except (TypeError, ValueError):
            pass
        cat = (p.get("category") or "Uncategorized").strip() or "Uncategorized"
        by_category[cat] = by_category.get(cat, 0) + 1
    return {
        "count": total,
        "annual_cost": annual,
        "licenses_purchased": licenses,
        "licenses_assigned": assigned,
        "by_category": sorted(by_category.items(), key=lambda kv: (-kv[1], kv[0])),
    }


@app.route("/engagements/<engagement_id>/inventory")
def engagement_inventory(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_inventory(eng):
        storage.save_engagement(eng)

    products = list(eng["inventory"]["products"])
    q = request.args.get("q", "").strip().lower()
    cat = request.args.get("category", "").strip()
    sort_key = request.args.get("sort", "product_name")
    sort_dir = request.args.get("dir", "asc")

    if q:
        def matches(p):
            blob = " ".join(str(v) for v in p.values()).lower()
            return q in blob
        products = [p for p in products if matches(p)]
    if cat:
        products = [p for p in products if (p.get("category") or "") == cat]

    def sort_value(p):
        v = p.get(sort_key, "")
        if sort_key in ("licenses_purchased", "licenses_assigned", "active_users"):
            try:
                return int(v) if v != "" else -1
            except (TypeError, ValueError):
                return -1
        if sort_key in ("cost_per_license", "total_annual_cost"):
            try:
                return float(v) if v != "" else -1.0
            except (TypeError, ValueError):
                return -1.0
        return (str(v) or "").lower()

    products.sort(key=sort_value, reverse=(sort_dir == "desc"))

    summary = _inventory_summary(eng["inventory"])
    return render_template(
        "inventory_list.html",
        eng=eng,
        products=products,
        summary=summary,
        categories=PRODUCT_CATEGORIES,
        sort_key=sort_key,
        sort_dir=sort_dir,
        q=q,
        category_filter=cat,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/inventory/new", methods=["GET", "POST"])
def engagement_inventory_new(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    if request.method == "POST":
        product = _new_product()
        _apply_form(product, request.form)
        if not product.get("product_name"):
            return render_template(
                "inventory_form.html",
                eng=eng, product=product, error="Product name is required.",
                categories=PRODUCT_CATEGORIES,
                deployment_models=DEPLOYMENT_MODELS,
                sensitivity_levels=DATA_SENSITIVITY_LEVELS,
                fields=PRODUCT_FIELDS,
                mode="new",
                phases=PHASES,
            )
        eng["inventory"]["products"].append(product)
        if eng["phase_progress"].get("inventory") == "not_started":
            eng["phase_progress"]["inventory"] = "in_progress"
        storage.save_engagement(eng)
        return redirect(url_for("engagement_inventory", engagement_id=engagement_id))

    return render_template(
        "inventory_form.html",
        eng=eng, product=_new_product(), error=None,
        categories=PRODUCT_CATEGORIES,
        deployment_models=DEPLOYMENT_MODELS,
        sensitivity_levels=DATA_SENSITIVITY_LEVELS,
        fields=PRODUCT_FIELDS,
        mode="new",
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/inventory/<product_id>/edit", methods=["GET", "POST"])
def engagement_inventory_edit(engagement_id, product_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    product = next((p for p in eng["inventory"]["products"] if p["id"] == product_id), None)
    if not product:
        abort(404)
    if request.method == "POST":
        _apply_form(product, request.form)
        if not product.get("product_name"):
            return render_template(
                "inventory_form.html",
                eng=eng, product=product, error="Product name is required.",
                categories=PRODUCT_CATEGORIES,
                deployment_models=DEPLOYMENT_MODELS,
                sensitivity_levels=DATA_SENSITIVITY_LEVELS,
                fields=PRODUCT_FIELDS,
                mode="edit",
                phases=PHASES,
            )
        storage.save_engagement(eng)
        return redirect(url_for("engagement_inventory", engagement_id=engagement_id))
    return render_template(
        "inventory_form.html",
        eng=eng, product=product, error=None,
        categories=PRODUCT_CATEGORIES,
        deployment_models=DEPLOYMENT_MODELS,
        sensitivity_levels=DATA_SENSITIVITY_LEVELS,
        fields=PRODUCT_FIELDS,
        mode="edit",
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/inventory/<product_id>/delete", methods=["POST"])
def engagement_inventory_delete(engagement_id, product_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    eng["inventory"]["products"] = [p for p in eng["inventory"]["products"] if p["id"] != product_id]
    storage.save_engagement(eng)
    return redirect(url_for("engagement_inventory", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/inventory/import", methods=["GET", "POST"])
def engagement_inventory_import(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    result = None
    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            result = {"error": "Choose a CSV or XLSX file to upload."}
        else:
            ext = os.path.splitext(f.filename)[1].lower()
            try:
                if ext == ".csv":
                    text = f.read().decode("utf-8-sig", errors="replace")
                    added, skipped, mapping = _import_csv_text(eng, text)
                elif ext in (".xlsx", ".xlsm"):
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(["" if c is None else str(c) for c in row])
                    sio = io.StringIO()
                    writer = csv.writer(sio)
                    writer.writerows(rows)
                    text = sio.getvalue()
                    added, skipped, mapping = _import_csv_text(eng, text)
                else:
                    result = {"error": f"Unsupported file type: {ext}. Use CSV or XLSX."}
                    added = skipped = 0
                    mapping = {}
            except Exception as ex:  # noqa: BLE001
                result = {"error": f"Import failed: {ex}"}
                added = skipped = 0
                mapping = {}
            if "error" not in (result or {}):
                if added > 0 and eng["phase_progress"].get("inventory") == "not_started":
                    eng["phase_progress"]["inventory"] = "in_progress"
                storage.save_engagement(eng)
                result = {"added": added, "skipped": skipped, "mapping": mapping}
    return render_template(
        "inventory_import.html",
        eng=eng, result=result,
        fields=PRODUCT_FIELDS,
        aliases=PRODUCT_FIELD_ALIASES,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/inventory/finalize", methods=["POST"])
def engagement_inventory_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["inventory"]["finalized"] = False
        eng["inventory"]["finalized_at"] = None
        eng["phase_progress"]["inventory"] = "in_progress"
        eng["status"] = "inventory"
    else:
        eng["inventory"]["finalized"] = True
        eng["inventory"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["inventory"] = "complete"
        eng["status"] = "normalize"
        if eng["phase_progress"].get("normalize") == "not_started":
            eng["phase_progress"]["normalize"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_inventory", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/inventory/export.csv")
def engagement_inventory_export_csv(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([label for _key, label, _t in PRODUCT_FIELDS])
    for p in eng["inventory"]["products"]:
        writer.writerow([p.get(key, "") for key, _label, _t in PRODUCT_FIELDS])
    fn = f"software-inventory-{eng.get('id')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.route("/engagements/<engagement_id>/inventory/export.xlsx")
def engagement_inventory_export_xlsx(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_inventory(eng)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Software Inventory"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B5FD9")
    headers = [label for _key, label, _t in PRODUCT_FIELDS]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for p in eng["inventory"]["products"]:
        row = []
        for key, _label, ftype in PRODUCT_FIELDS:
            v = p.get(key, "")
            if ftype in ("int",) and v != "":
                try:
                    v = int(v)
                except (TypeError, ValueError):
                    pass
            elif ftype == "money" and v != "":
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    pass
            row.append(v)
        ws.append(row)

    widths = [22, 18, 10, 24, 18, 18, 18, 14, 14, 12, 14, 16, 14, 16, 18, 22, 24, 24, 16, 26, 28, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fn = f"software-inventory-{eng.get('id')}.xlsx"
    return Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.route("/engagements/<engagement_id>/inventory/template.csv")
def engagement_inventory_template_csv(engagement_id):
    """Empty CSV with the canonical headers — useful for customers preparing data."""
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([label for _key, label, _t in PRODUCT_FIELDS])
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="software-inventory-template.csv"'},
    )


@app.template_filter("format_bytes")
def format_bytes(n):
    if n is None:
        return ""
    n = float(n)
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{int(n)} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


# ---------------------------------------------------------------------------
# Phase 4 — Normalize the Data
# ---------------------------------------------------------------------------


def _ensure_normalize(eng):
    if "normalize" in eng and "ignored_issues" in eng["normalize"]:
        return False
    eng["normalize"] = {
        "ignored_issues": {},  # issue_id -> {"ignored_at": iso, "reason": str}
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _norm_text(s):
    return (s or "").strip().lower().replace("  ", " ")


_VENDOR_SUFFIXES = (
    " incorporated", " corporation", " technologies", " technology",
    ", inc", ", inc.", " inc.", " inc",
    ", corp", ", corp.", " corp.", " corp",
    " llc", " l.l.c.", " ltd.", " ltd",
    " gmbh", " plc", " ag", " sa", " spa", " bv", " nv",
    " co.", " co",
    " (us)", " (uk)", " (eu)",
)


def _norm_vendor(s):
    """Aggressive vendor normalization: lowercase, trim corporate suffixes & punctuation."""
    s = _norm_text(s)
    if not s:
        return ""
    changed = True
    while changed:
        changed = False
        for sfx in _VENDOR_SUFFIXES:
            if s.endswith(sfx):
                s = s[:-len(sfx)].rstrip(", ")
                changed = True
    return s.strip().rstrip(",.")


def detect_duplicates(products):
    """Group products with same normalized (name, vendor); return clusters of 2+."""
    groups = {}
    for p in products:
        name = _norm_text(p.get("product_name"))
        vendor = _norm_text(p.get("vendor"))
        if not name:
            continue
        groups.setdefault((name, vendor), []).append(p)
    out = []
    for (name, vendor), members in groups.items():
        if len(members) >= 2:
            ids = sorted(m["id"] for m in members)
            issue_id = "dup:" + ":".join(ids)
            out.append({
                "issue_id": issue_id,
                "label": members[0].get("product_name") + (f" — {members[0].get('vendor')}" if members[0].get("vendor") else ""),
                "members": members,
            })
    return out


def detect_vendor_variants(products, threshold=0.85):
    """
    Two passes:
      1. Group by aggressively-normalized vendor (suffix-stripped). Anything
         with 2+ raw spellings under one normalized key is a cluster.
      2. difflib over the remaining single-spelling normalized vendors to catch
         typos (e.g. "Atlasian" vs "Atlassian").
    """
    norm_to_originals = {}      # vendor_norm (suffix-stripped) -> set of raw vendor strings
    norm_to_text_norms = {}     # vendor_norm -> set of plain text-norm strings (for cluster_norms)
    for p in products:
        raw = p.get("vendor")
        if not raw:
            continue
        vn = _norm_vendor(raw)
        if not vn:
            continue
        norm_to_originals.setdefault(vn, set()).add(raw)
        norm_to_text_norms.setdefault(vn, set()).add(_norm_text(raw))

    clusters = []
    handled_text_norms = set()

    # pass 1: same suffix-stripped vendor, multiple spellings
    for vn, originals in norm_to_originals.items():
        if len(originals) >= 2:
            text_norms = sorted(norm_to_text_norms[vn])
            handled_text_norms.update(text_norms)
            cluster = _build_vendor_cluster(products, sorted(originals), text_norms)
            clusters.append(cluster)

    # pass 2: difflib over the suffix-stripped names (catches typos)
    remaining = [vn for vn, origs in norm_to_originals.items() if len(origs) == 1]
    seen = set()
    for v in remaining:
        if v in seen:
            continue
        matches = difflib.get_close_matches(v, remaining, n=10, cutoff=threshold)
        matches = [m for m in matches if m != v]
        if matches:
            group = sorted(set([v] + matches))
            seen.update(group)
            originals = sorted({orig for vn in group for orig in norm_to_originals[vn]})
            text_norms = sorted({tn for vn in group for tn in norm_to_text_norms[vn]})
            if any(tn in handled_text_norms for tn in text_norms):
                continue
            handled_text_norms.update(text_norms)
            clusters.append(_build_vendor_cluster(products, originals, text_norms))

    return clusters


def _build_vendor_cluster(products, originals, cluster_norms):
    counts = {}
    affected = []
    for p in products:
        if _norm_text(p.get("vendor")) in cluster_norms:
            counts[p["vendor"]] = counts.get(p["vendor"], 0) + 1
            affected.append(p)
    canonical = max(counts.items(), key=lambda kv: (kv[1], len(kv[0])))[0] if counts else originals[0]
    return {
        "issue_id": "vv:" + "|".join(sorted(cluster_norms)),
        "variants": originals,
        "canonical_suggested": canonical,
        "affected_count": len(affected),
        "affected_products": affected,
        "cluster_norms": sorted(cluster_norms),
    }


def detect_unmapped_categories(products):
    """Categories that don't match PRODUCT_CATEGORIES (typos / customer-supplied)."""
    valid = set(PRODUCT_CATEGORIES)
    out = {}
    for p in products:
        cat = (p.get("category") or "").strip()
        if cat and cat not in valid:
            out.setdefault(cat, []).append(p)
    items = []
    for cat, members in out.items():
        # propose closest valid category
        suggestions = difflib.get_close_matches(cat, PRODUCT_CATEGORIES, n=1, cutoff=0.6)
        items.append({
            "issue_id": f"cv:{cat}",
            "category": cat,
            "members": members,
            "suggested": suggestions[0] if suggestions else "",
        })
    return items


def detect_missing_owners(products):
    out = []
    for p in products:
        missing = []
        if not (p.get("business_owner") or "").strip():
            missing.append("business")
        if not (p.get("technical_owner") or "").strip():
            missing.append("technical")
        if not (p.get("contract_owner") or "").strip():
            missing.append("contract")
        if missing:
            out.append({
                "issue_id": f"mo:{p['id']}:{','.join(missing)}",
                "product": p,
                "missing": missing,
            })
    return out


def detect_missing_cost(products):
    out = []
    for p in products:
        if not (p.get("total_annual_cost") or p.get("cost_per_license")):
            out.append({"issue_id": f"mc:{p['id']}", "product": p})
    return out


def detect_uncategorized(products):
    out = []
    for p in products:
        cat = (p.get("category") or "").strip()
        if not cat:
            out.append({"issue_id": f"uc:{p['id']}", "product": p})
    return out


def detect_unclear_use(products):
    out = []
    for p in products:
        if not (p.get("primary_use_case") or "").strip():
            out.append({"issue_id": f"mu:{p['id']}", "product": p})
    return out


def detect_license_anomalies(products):
    out = []
    for p in products:
        def to_int(v):
            try:
                return int(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None
        purch = to_int(p.get("licenses_purchased"))
        asgn = to_int(p.get("licenses_assigned"))
        users = to_int(p.get("active_users"))
        if purch is not None and asgn is not None and asgn > purch:
            out.append({
                "issue_id": f"la:{p['id']}:assigned_over_purchased",
                "product": p,
                "type": "Assigned exceeds purchased",
                "detail": f"{asgn} assigned vs {purch} purchased",
            })
        if asgn is not None and users is not None and users > asgn:
            out.append({
                "issue_id": f"la:{p['id']}:users_over_assigned",
                "product": p,
                "type": "Active users exceed assigned licenses",
                "detail": f"{users} active users vs {asgn} assigned",
            })
    return out


def collect_normalize_findings(eng):
    """Run all detectors. Filter out ignored issues. Return a structured report."""
    products = eng.get("inventory", {}).get("products", [])
    ignored = eng.get("normalize", {}).get("ignored_issues", {})

    def keep(issues):
        return [i for i in issues if i["issue_id"] not in ignored]

    findings = {
        "duplicates": keep(detect_duplicates(products)),
        "vendor_variants": keep(detect_vendor_variants(products)),
        "unmapped_categories": keep(detect_unmapped_categories(products)),
        "uncategorized": keep(detect_uncategorized(products)),
        "missing_owners": keep(detect_missing_owners(products)),
        "missing_cost": keep(detect_missing_cost(products)),
        "unclear_use": keep(detect_unclear_use(products)),
        "license_anomalies": keep(detect_license_anomalies(products)),
    }
    findings["total_open"] = sum(len(v) for k, v in findings.items() if k != "total_open")
    findings["ignored_count"] = len(ignored)
    return findings


@app.route("/engagements/<engagement_id>/normalize")
def engagement_normalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_normalize(eng):
        if (eng["phase_progress"].get("inventory") == "complete"
                and eng["phase_progress"].get("normalize") == "not_started"):
            eng["phase_progress"]["normalize"] = "in_progress"
        storage.save_engagement(eng)
    findings = collect_normalize_findings(eng)
    return render_template(
        "normalize.html",
        eng=eng,
        findings=findings,
        categories=PRODUCT_CATEGORIES,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/normalize/apply-vendor", methods=["POST"])
def engagement_normalize_apply_vendor(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_normalize(eng)
    cluster_norms_csv = request.form.get("cluster_norms", "")
    canonical = request.form.get("canonical", "").strip()
    if not canonical or not cluster_norms_csv:
        return redirect(url_for("engagement_normalize", engagement_id=engagement_id))
    cluster_norms = set(filter(None, cluster_norms_csv.split("|")))
    changed = 0
    now = datetime.utcnow().isoformat() + "Z"
    for p in eng["inventory"]["products"]:
        if _norm_text(p.get("vendor")) in cluster_norms and p.get("vendor") != canonical:
            p["vendor"] = canonical
            p["updated_at"] = now
            changed += 1
    storage.save_engagement(eng)
    return redirect(url_for("engagement_normalize", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/normalize/apply-category", methods=["POST"])
def engagement_normalize_apply_category(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_normalize(eng)
    old_category = request.form.get("old_category", "").strip()
    new_category = request.form.get("new_category", "").strip()
    if not old_category or not new_category:
        return redirect(url_for("engagement_normalize", engagement_id=engagement_id))
    if new_category not in PRODUCT_CATEGORIES:
        # allow free-text; just trust the user
        pass
    now = datetime.utcnow().isoformat() + "Z"
    for p in eng["inventory"]["products"]:
        if (p.get("category") or "").strip() == old_category:
            p["category"] = new_category
            p["updated_at"] = now
    storage.save_engagement(eng)
    return redirect(url_for("engagement_normalize", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/normalize/merge-duplicates", methods=["POST"])
def engagement_normalize_merge_duplicates(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_normalize(eng)
    keep_id = request.form.get("keep_id", "").strip()
    delete_ids_csv = request.form.get("delete_ids", "")
    delete_ids = set(filter(None, delete_ids_csv.split(",")))
    if not keep_id or not delete_ids:
        return redirect(url_for("engagement_normalize", engagement_id=engagement_id))
    eng["inventory"]["products"] = [p for p in eng["inventory"]["products"] if p["id"] not in delete_ids]
    storage.save_engagement(eng)
    return redirect(url_for("engagement_normalize", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/normalize/ignore", methods=["POST"])
def engagement_normalize_ignore(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_normalize(eng)
    issue_id = request.form.get("issue_id", "").strip()
    reason = request.form.get("reason", "").strip()
    action = request.form.get("action", "ignore")
    if not issue_id:
        return redirect(url_for("engagement_normalize", engagement_id=engagement_id))
    if action == "unignore":
        eng["normalize"]["ignored_issues"].pop(issue_id, None)
    else:
        eng["normalize"]["ignored_issues"][issue_id] = {
            "ignored_at": datetime.utcnow().isoformat() + "Z",
            "reason": reason,
        }
    storage.save_engagement(eng)
    return redirect(url_for("engagement_normalize", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/normalize/finalize", methods=["POST"])
def engagement_normalize_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_normalize(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["normalize"]["finalized"] = False
        eng["normalize"]["finalized_at"] = None
        eng["phase_progress"]["normalize"] = "in_progress"
        eng["status"] = "normalize"
    else:
        eng["normalize"]["finalized"] = True
        eng["normalize"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["normalize"] = "complete"
        eng["status"] = "overlap"
        if eng["phase_progress"].get("overlap") == "not_started":
            eng["phase_progress"]["overlap"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_normalize", engagement_id=engagement_id))


# ---------------------------------------------------------------------------
# Phase 5 — Identify Product Overlap
# ---------------------------------------------------------------------------

def _ensure_overlap(eng):
    if "overlap" in eng and "dispositions" in eng["overlap"]:
        return False
    eng["overlap"] = {
        "dispositions": {},  # product_id -> {disposition, risk_of_removal, notes, updated_at}
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _safe_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _safe_int(v):
    try:
        return int(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


def _overlap_clusters(products):
    """Group products by category. Categories with 2+ products are overlap clusters."""
    groups = {}
    for p in products:
        cat = (p.get("category") or "Uncategorized").strip() or "Uncategorized"
        groups.setdefault(cat, []).append(p)
    clusters = []
    for cat, members in sorted(groups.items()):
        if len(members) >= 2:
            annual = sum(_safe_float(p.get("total_annual_cost")) for p in members)
            users = sum(_safe_int(p.get("active_users")) for p in members)
            licenses = sum(_safe_int(p.get("licenses_purchased")) for p in members)
            clusters.append({
                "category": cat,
                "products": members,
                "annual_total": annual,
                "users_total": users,
                "licenses_total": licenses,
            })
    return clusters


def _overlap_summary(eng):
    products = eng.get("inventory", {}).get("products", [])
    clusters = _overlap_clusters(products)
    overlap_pids = {p["id"] for c in clusters for p in c["products"]}
    annual_in_overlap = 0.0
    for p in products:
        if p["id"] in overlap_pids:
            annual_in_overlap += _safe_float(p.get("total_annual_cost"))

    dispositions = eng.get("overlap", {}).get("dispositions", {})
    potential_savings = 0.0
    decided = 0
    products_by_id = {p["id"]: p for p in products}
    for pid in overlap_pids:
        d = dispositions.get(pid, {}).get("disposition", "")
        if d:
            decided += 1
        if d in SAVINGS_DISPOSITIONS and pid in products_by_id:
            potential_savings += _safe_float(products_by_id[pid].get("total_annual_cost"))

    return {
        "category_count": len(clusters),
        "products_in_overlap": len(overlap_pids),
        "annual_in_overlap": annual_in_overlap,
        "potential_savings": potential_savings,
        "decided_count": decided,
        "undecided_count": len(overlap_pids) - decided,
    }


@app.route("/engagements/<engagement_id>/overlap", methods=["GET", "POST"])
def engagement_overlap(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_overlap(eng):
        if (eng["phase_progress"].get("normalize") == "complete"
                and eng["phase_progress"].get("overlap") == "not_started"):
            eng["phase_progress"]["overlap"] = "in_progress"
        storage.save_engagement(eng)

    if request.method == "POST":
        product_ids = request.form.getlist("product_id")
        valid_dispositions = dict(DISPOSITIONS)
        valid_risks = dict(REMOVAL_RISKS)
        now = datetime.utcnow().isoformat() + "Z"
        for pid in product_ids:
            disp = request.form.get(f"disposition_{pid}", "").strip()
            risk = request.form.get(f"risk_{pid}", "").strip()
            notes = request.form.get(f"notes_{pid}", "").strip()
            if disp and disp not in valid_dispositions:
                disp = ""
            if risk and risk not in valid_risks:
                risk = ""
            if not (disp or risk or notes):
                eng["overlap"]["dispositions"].pop(pid, None)
                continue
            eng["overlap"]["dispositions"][pid] = {
                "disposition": disp,
                "risk_of_removal": risk,
                "notes": notes,
                "updated_at": now,
            }
        if eng["phase_progress"].get("overlap") == "not_started":
            eng["phase_progress"]["overlap"] = "in_progress"
        storage.save_engagement(eng)
        return redirect(url_for("engagement_overlap", engagement_id=engagement_id))

    products = eng.get("inventory", {}).get("products", [])
    clusters = _overlap_clusters(products)
    summary = _overlap_summary(eng)
    return render_template(
        "overlap.html",
        eng=eng,
        clusters=clusters,
        summary=summary,
        dispositions=eng["overlap"]["dispositions"],
        disposition_options=DISPOSITIONS,
        risk_options=REMOVAL_RISKS,
        savings_dispositions=SAVINGS_DISPOSITIONS,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/overlap/finalize", methods=["POST"])
def engagement_overlap_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_overlap(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["overlap"]["finalized"] = False
        eng["overlap"]["finalized_at"] = None
        eng["phase_progress"]["overlap"] = "in_progress"
        eng["status"] = "overlap"
    else:
        eng["overlap"]["finalized"] = True
        eng["overlap"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["overlap"] = "complete"
        eng["status"] = "ai_review"
        if eng["phase_progress"].get("ai_review") == "not_started":
            eng["phase_progress"]["ai_review"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_overlap", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/overlap/analysis")
def engagement_overlap_analysis(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_overlap(eng)
    products = eng.get("inventory", {}).get("products", [])
    clusters = _overlap_clusters(products)
    summary = _overlap_summary(eng)
    return render_template(
        "overlap_analysis.html",
        eng=eng,
        clusters=clusters,
        summary=summary,
        dispositions=eng["overlap"]["dispositions"],
        disposition_label=dict(DISPOSITIONS),
        risk_label=dict(REMOVAL_RISKS),
        savings_dispositions=SAVINGS_DISPOSITIONS,
    )


@app.route("/engagements/<engagement_id>/overlap/analysis.csv")
def engagement_overlap_analysis_csv(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_overlap(eng)
    products = eng.get("inventory", {}).get("products", [])
    clusters = _overlap_clusters(products)
    dispositions = eng["overlap"]["dispositions"]
    disposition_label = dict(DISPOSITIONS)
    risk_label = dict(REMOVAL_RISKS)

    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([
        "Category", "Product", "Vendor", "Active users", "Licenses purchased",
        "Licenses assigned", "Annual cost", "Renewal date", "Contract term",
        "Disposition", "Risk of removal", "Notes",
    ])
    for c in clusters:
        for p in c["products"]:
            d = dispositions.get(p["id"], {})
            writer.writerow([
                c["category"],
                p.get("product_name", ""),
                p.get("vendor", ""),
                p.get("active_users", ""),
                p.get("licenses_purchased", ""),
                p.get("licenses_assigned", ""),
                p.get("total_annual_cost", ""),
                p.get("renewal_date", ""),
                p.get("contract_term", ""),
                disposition_label.get(d.get("disposition", ""), ""),
                risk_label.get(d.get("risk_of_removal", ""), ""),
                d.get("notes", ""),
            ])
    fn = f"product-overlap-analysis-{eng.get('id')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 9 — Stakeholder Validation
# ---------------------------------------------------------------------------

def _ensure_validation(eng):
    if "validation" in eng and "validations" in eng["validation"]:
        return False
    eng["validation"] = {
        "validations": {},
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _new_validation_record():
    return {
        "stakeholders": [],
        "answers": {key: "" for key, _label in VALIDATION_QUESTIONS},
        "overall_status": "not_started",
        "notes": "",
        "updated_at": "",
    }


def _parse_stakeholders(text):
    """Parse 'Name | Role | YYYY-MM-DD | Status | notes' lines into records."""
    role_lookup_label = {l.lower(): k for k, l in STAKEHOLDER_ROLES}
    role_lookup_code = {k: l for k, l in STAKEHOLDER_ROLES}
    status_lookup_label = {l.lower(): k for k, l in STAKEHOLDER_STATUSES}
    status_lookup_code = {k: l for k, l in STAKEHOLDER_STATUSES}
    out = []
    for line in _parse_lines(text):
        parts = [p.strip() for p in line.split("|")]
        if not parts or not parts[0]:
            continue
        rec = {
            "name": parts[0],
            "role_code": "",
            "role_label": "",
            "consulted_date": "",
            "status_code": "",
            "status_label": "",
            "notes": "",
        }
        if len(parts) > 1 and parts[1]:
            raw = parts[1]
            low = raw.lower()
            if low in role_lookup_label:
                rec["role_code"] = role_lookup_label[low]
                rec["role_label"] = role_lookup_code[rec["role_code"]]
            elif low in role_lookup_code:
                rec["role_code"] = low
                rec["role_label"] = role_lookup_code[low]
            else:
                rec["role_label"] = raw
        if len(parts) > 2 and parts[2]:
            rec["consulted_date"] = _coerce_date(parts[2])
        if len(parts) > 3 and parts[3]:
            raw = parts[3]
            low = raw.lower()
            if low in status_lookup_label:
                rec["status_code"] = status_lookup_label[low]
                rec["status_label"] = status_lookup_code[rec["status_code"]]
            elif low in status_lookup_code:
                rec["status_code"] = low
                rec["status_label"] = status_lookup_code[low]
            else:
                rec["status_label"] = raw
        if len(parts) > 4:
            rec["notes"] = "|".join(parts[4:]).strip()
        out.append(rec)
    return out


def _stakeholders_to_text(stakeholders):
    lines = []
    for s in stakeholders or []:
        bits = [s.get("name", "")]
        if s.get("role_label"):
            bits.append(s["role_label"])
        if s.get("consulted_date"):
            bits.append(s["consulted_date"])
        if s.get("status_label"):
            bits.append(s["status_label"])
        if s.get("notes"):
            bits.append(s["notes"])
        lines.append(" | ".join(bits))
    return "\n".join(lines)


def _validation_summary(eng):
    opps = eng.get("savings", {}).get("opportunities", {})
    val_records = eng.get("validation", {}).get("validations", {})
    by_status = {"not_started": 0, "pending": 0, "validated": 0, "blocked": 0}
    total_with_validation = 0
    total_stakeholders = 0
    for opp_id in opps:
        rec = val_records.get(opp_id)
        st = (rec or {}).get("overall_status", "not_started")
        if st in by_status:
            by_status[st] += 1
        else:
            by_status["not_started"] += 1
        if rec and (rec.get("stakeholders") or any(rec.get("answers", {}).values())):
            total_with_validation += 1
        if rec:
            total_stakeholders += len(rec.get("stakeholders", []))
    return {
        "opportunity_count": len(opps),
        "by_status": by_status,
        "validations_started": total_with_validation,
        "stakeholders_recorded": total_stakeholders,
    }


@app.route("/engagements/<engagement_id>/validation", methods=["GET"])
def engagement_validation(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_validation(eng):
        if (eng["phase_progress"].get("savings") == "complete"
                and eng["phase_progress"].get("validation") == "not_started"):
            eng["phase_progress"]["validation"] = "in_progress"
        storage.save_engagement(eng)

    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    opps = eng.get("savings", {}).get("opportunities", {})

    # Sort: not-yet-validated first, then by status order, then by recurring savings desc.
    val_records = eng["validation"]["validations"]
    status_order = {"blocked": 0, "pending": 1, "not_started": 2, "validated": 3}
    opps_sorted = sorted(
        opps.values(),
        key=lambda o: (
            status_order.get((val_records.get(o["id"]) or {}).get("overall_status", "not_started"), 4),
            -float(o.get("recurring_annual_savings") or 0),
        ),
    )
    annotated = []
    for o in opps_sorted:
        details = []
        for pid in o.get("product_ids", []):
            p = products_by_id.get(pid)
            if p:
                details.append({
                    "product_name": p.get("product_name", ""),
                    "vendor": p.get("vendor", ""),
                    "category": p.get("category", ""),
                })
        rec = val_records.get(o["id"]) or _new_validation_record()
        annotated.append({
            **o,
            "products": details,
            "totals": _opportunity_totals(o),
            "validation": rec,
            "stakeholders_text": _stakeholders_to_text(rec.get("stakeholders", [])),
        })

    summary = _validation_summary(eng)
    return render_template(
        "validation.html",
        eng=eng,
        opportunities=annotated,
        summary=summary,
        questions=VALIDATION_QUESTIONS,
        roles=STAKEHOLDER_ROLES,
        statuses=STAKEHOLDER_STATUSES,
        overall_statuses=VALIDATION_OVERALL_STATUSES,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/validation/<opp_id>", methods=["POST"])
def engagement_validation_save(engagement_id, opp_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_validation(eng)
    if opp_id not in eng.get("savings", {}).get("opportunities", {}):
        abort(404)

    rec = eng["validation"]["validations"].get(opp_id) or _new_validation_record()
    rec["stakeholders"] = _parse_stakeholders(request.form.get("stakeholders", ""))
    answers = {}
    for key, _label in VALIDATION_QUESTIONS:
        answers[key] = request.form.get(f"answer_{key}", "").strip()
    rec["answers"] = answers
    overall = request.form.get("overall_status", "").strip()
    valid_overall = {k for k, _l in VALIDATION_OVERALL_STATUSES}
    if overall in valid_overall:
        rec["overall_status"] = overall
    rec["notes"] = request.form.get("notes", "").strip()
    rec["updated_at"] = datetime.utcnow().isoformat() + "Z"

    # Empty record (no stakeholders, no answers, no overall, no notes) -> drop entirely
    is_empty = (
        not rec["stakeholders"]
        and not any(answers.values())
        and rec["overall_status"] == "not_started"
        and not rec["notes"]
    )
    if is_empty:
        eng["validation"]["validations"].pop(opp_id, None)
    else:
        eng["validation"]["validations"][opp_id] = rec

    if eng["phase_progress"].get("validation") == "not_started":
        eng["phase_progress"]["validation"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_validation", engagement_id=engagement_id) + f"#opp-{opp_id}")


@app.route("/engagements/<engagement_id>/validation/finalize", methods=["POST"])
def engagement_validation_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_validation(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["validation"]["finalized"] = False
        eng["validation"]["finalized_at"] = None
        eng["phase_progress"]["validation"] = "in_progress"
        eng["status"] = "validation"
    else:
        eng["validation"]["finalized"] = True
        eng["validation"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["validation"] = "complete"
        eng["status"] = "recommendations"
        if eng["phase_progress"].get("recommendations") == "not_started":
            eng["phase_progress"]["recommendations"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_validation", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/validation/notes")
def engagement_validation_notes(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_validation(eng)
    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    opps = eng.get("savings", {}).get("opportunities", {})
    val_records = eng["validation"]["validations"]
    status_order = {"validated": 0, "pending": 1, "blocked": 2, "not_started": 3}
    opps_sorted = sorted(
        opps.values(),
        key=lambda o: (
            status_order.get((val_records.get(o["id"]) or {}).get("overall_status", "not_started"), 4),
            o.get("title", ""),
        ),
    )
    annotated = []
    for o in opps_sorted:
        rec = val_records.get(o["id"])
        if not rec:
            continue
        if not (rec.get("stakeholders") or any(rec.get("answers", {}).values()) or rec.get("notes")):
            continue
        details = [products_by_id.get(pid, {}).get("product_name", "")
                   for pid in o.get("product_ids", []) if pid in products_by_id]
        annotated.append({**o, "validation": rec, "product_names": details})
    summary = _validation_summary(eng)
    return render_template(
        "validation_notes.html",
        eng=eng,
        opportunities=annotated,
        summary=summary,
        questions=VALIDATION_QUESTIONS,
        overall_status_label={k: l for k, l in VALIDATION_OVERALL_STATUSES},
    )


@app.route("/engagements/<engagement_id>/validation/notes.csv")
def engagement_validation_notes_csv(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_validation(eng)
    products_by_id = {p["id"]: p for p in eng.get("inventory", {}).get("products", [])}
    opps = eng.get("savings", {}).get("opportunities", {})
    val_records = eng["validation"]["validations"]
    overall_status_label = {k: l for k, l in VALIDATION_OVERALL_STATUSES}

    sio = io.StringIO()
    writer = csv.writer(sio)
    header = [
        "Opportunity", "Products", "Overall status",
        "Stakeholders consulted", "Stakeholders agreed",
        "Stakeholders pushback", "Stakeholders blocked",
    ]
    for _key, label in VALIDATION_QUESTIONS:
        header.append(label)
    header.append("Notes")
    writer.writerow(header)

    for opp_id, rec in val_records.items():
        opp = opps.get(opp_id)
        if not opp:
            continue
        names = [products_by_id.get(pid, {}).get("product_name", "")
                 for pid in opp.get("product_ids", []) if pid in products_by_id]
        sh = rec.get("stakeholders", [])
        counts = {"consulted": 0, "agreed": 0, "pushback": 0, "blocked": 0}
        for s in sh:
            sc = s.get("status_code") or ""
            if sc in counts:
                counts[sc] += 1
        row = [
            opp.get("title", ""),
            "; ".join(names),
            overall_status_label.get(rec.get("overall_status", ""), rec.get("overall_status", "")),
            counts["consulted"], counts["agreed"], counts["pushback"], counts["blocked"],
        ]
        for key, _label in VALIDATION_QUESTIONS:
            row.append(rec.get("answers", {}).get(key, ""))
        row.append(rec.get("notes", ""))
        writer.writerow(row)

    fn = f"stakeholder-validation-notes-{eng.get('id')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 8 — Estimate Cost Savings
# ---------------------------------------------------------------------------

def _ensure_savings(eng):
    if "savings" in eng and "opportunities" in eng["savings"]:
        return False
    eng["savings"] = {
        "opportunities": {},  # opp_id -> dict
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _new_opportunity(*, title, source, seed_key, product_ids, disposition,
                     current_annual_cost, recurring_annual_savings,
                     migration_cost=0.0, training_cost=0.0, one_time_savings=0.0,
                     notes="", status="proposed"):
    oid = uuid.uuid4().hex[:8]
    now = datetime.utcnow().isoformat() + "Z"
    return {
        "id": oid,
        "title": title,
        "source": source,                 # "phase5" | "phase7-unused" | "manual"
        "seed_key": seed_key,             # stable key to avoid duplicate seeding
        "product_ids": list(product_ids),
        "disposition": disposition,
        "current_annual_cost": float(current_annual_cost),
        "recurring_annual_savings": float(recurring_annual_savings),
        "migration_cost": float(migration_cost),
        "training_cost": float(training_cost),
        "one_time_savings": float(one_time_savings),
        "notes": notes,
        "status": status,
        "created_at": now,
        "updated_at": now,
    }


def _opportunity_totals(opp):
    """Compute display values: first-year net savings + recurring."""
    rec = float(opp.get("recurring_annual_savings") or 0)
    one = float(opp.get("one_time_savings") or 0)
    mig = float(opp.get("migration_cost") or 0)
    trn = float(opp.get("training_cost") or 0)
    return {
        "first_year": rec + one - mig - trn,
        "recurring": rec,
        "transition_total": mig + trn,
    }


def _seed_savings_opportunities(eng):
    """
    Seed opportunities from:
      1. Phase 5 dispositions (retire / replace / consolidate / renegotiate)
      2. Phase 7 'unused' flags that aren't already covered by a Phase 5 entry
    Returns the number of new opportunities created.
    """
    _ensure_savings(eng)
    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    existing_seeds = {
        opp.get("seed_key")
        for opp in eng["savings"]["opportunities"].values()
        if opp.get("seed_key")
    }
    created = 0

    # 1. Phase 5 dispositions
    p5_dispositions = eng.get("overlap", {}).get("dispositions", {})
    seeded_from_p5 = set()
    for pid, drec in p5_dispositions.items():
        disp = (drec.get("disposition") or "").strip()
        if disp not in {"retire", "replace", "consolidate", "renegotiate"}:
            continue
        seed_key = f"phase5:{pid}"
        seeded_from_p5.add(pid)
        if seed_key in existing_seeds:
            continue
        p = products_by_id.get(pid)
        if not p:
            continue
        annual = _safe_float(p.get("total_annual_cost"))
        # Renegotiate doesn't capture the whole cost — default to 20% recurring.
        recurring = annual * (0.2 if disp == "renegotiate" else 1.0)
        title = f"{disp.title()} {p.get('product_name', 'product')}"
        if disp == "consolidate" and p.get("category"):
            title += f" — {p['category']}"
        opp = _new_opportunity(
            title=title,
            source="phase5",
            seed_key=seed_key,
            product_ids=[pid],
            disposition=disp,
            current_annual_cost=annual,
            recurring_annual_savings=recurring,
            notes=(drec.get("notes") or "").strip(),
        )
        eng["savings"]["opportunities"][opp["id"]] = opp
        existing_seeds.add(seed_key)
        created += 1

    # 2. Phase 7 'unused' flags not already covered by Phase 5
    p7_flags = eng.get("tech_debt", {}).get("flags", {})
    for pid, frec in p7_flags.items():
        if "unused" not in (frec.get("flags") or []):
            continue
        if pid in seeded_from_p5:
            continue  # Phase 5 retire/replace already covers it
        seed_key = f"phase7-unused:{pid}"
        if seed_key in existing_seeds:
            continue
        p = products_by_id.get(pid)
        if not p:
            continue
        annual = _safe_float(p.get("total_annual_cost"))
        purchased = _safe_int(p.get("licenses_purchased"))
        active = _safe_int(p.get("active_users"))
        cpl = _safe_float(p.get("cost_per_license"))
        # Recurring = cost_per_license * unused-license-count.
        # Fallback: if no per-license cost, assume reducing to active users.
        if cpl > 0 and purchased > 0:
            unused = max(purchased - active, 0)
            recurring = cpl * unused
        elif purchased > 0:
            unused_ratio = max(purchased - active, 0) / purchased
            recurring = annual * unused_ratio
        else:
            recurring = annual
        opp = _new_opportunity(
            title=f"Reduce unused licenses on {p.get('product_name', 'product')}",
            source="phase7-unused",
            seed_key=seed_key,
            product_ids=[pid],
            disposition="reduce_licenses",
            current_annual_cost=annual,
            recurring_annual_savings=recurring,
            notes=(frec.get("notes") or "").strip(),
        )
        eng["savings"]["opportunities"][opp["id"]] = opp
        existing_seeds.add(seed_key)
        created += 1

    return created


def _savings_summary(eng):
    opps = eng.get("savings", {}).get("opportunities", {}).values()
    by_status = {"proposed": 0, "approved": 0, "rejected": 0}
    total_recurring = 0.0
    total_first_year = 0.0
    total_transition = 0.0
    total_current_cost = 0.0
    approved_recurring = 0.0
    approved_first_year = 0.0
    for opp in opps:
        st = opp.get("status", "proposed")
        if st in by_status:
            by_status[st] += 1
        if st == "rejected":
            continue
        t = _opportunity_totals(opp)
        total_recurring += t["recurring"]
        total_first_year += t["first_year"]
        total_transition += t["transition_total"]
        total_current_cost += float(opp.get("current_annual_cost") or 0)
        if st == "approved":
            approved_recurring += t["recurring"]
            approved_first_year += t["first_year"]
    return {
        "count": len(eng.get("savings", {}).get("opportunities", {})),
        "by_status": by_status,
        "total_current_cost": total_current_cost,
        "total_recurring": total_recurring,
        "total_first_year": total_first_year,
        "total_transition": total_transition,
        "approved_recurring": approved_recurring,
        "approved_first_year": approved_first_year,
    }


@app.route("/engagements/<engagement_id>/savings", methods=["GET"])
def engagement_savings(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_savings(eng):
        if (eng["phase_progress"].get("tech_debt") == "complete"
                and eng["phase_progress"].get("savings") == "not_started"):
            eng["phase_progress"]["savings"] = "in_progress"
        storage.save_engagement(eng)

    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    opps_sorted = sorted(
        eng["savings"]["opportunities"].values(),
        key=lambda o: (
            {"approved": 0, "proposed": 1, "rejected": 2}.get(o.get("status"), 3),
            -float(o.get("recurring_annual_savings") or 0),
        ),
    )
    # Annotate each opp with computed totals + product details for the template.
    annotated = []
    for o in opps_sorted:
        details = []
        for pid in o.get("product_ids", []):
            p = products_by_id.get(pid)
            if p:
                details.append({
                    "id": pid,
                    "product_name": p.get("product_name", ""),
                    "vendor": p.get("vendor", ""),
                    "category": p.get("category", ""),
                    "annual_cost": _safe_float(p.get("total_annual_cost")),
                })
        annotated.append({**o, "totals": _opportunity_totals(o), "products": details})

    summary = _savings_summary(eng)
    return render_template(
        "savings.html",
        eng=eng,
        opportunities=annotated,
        products=products,
        summary=summary,
        disposition_choices=SAVINGS_DISPOSITION_CHOICES,
        statuses=SAVINGS_STATUSES,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/savings/seed", methods=["POST"])
def engagement_savings_seed(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    created = _seed_savings_opportunities(eng)
    if eng["phase_progress"].get("savings") == "not_started":
        eng["phase_progress"]["savings"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id) + f"?seeded={created}")


@app.route("/engagements/<engagement_id>/savings/new", methods=["POST"])
def engagement_savings_new(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    title = request.form.get("title", "").strip() or "Untitled opportunity"
    product_ids = request.form.getlist("product_ids")
    disposition = request.form.get("disposition", "other").strip()
    valid_dispositions = {k for k, _l in SAVINGS_DISPOSITION_CHOICES}
    if disposition not in valid_dispositions:
        disposition = "other"
    products_by_id = {p["id"]: p for p in eng.get("inventory", {}).get("products", [])}
    current_cost = sum(_safe_float(products_by_id[pid].get("total_annual_cost"))
                       for pid in product_ids if pid in products_by_id)
    opp = _new_opportunity(
        title=title,
        source="manual",
        seed_key="",
        product_ids=product_ids,
        disposition=disposition,
        current_annual_cost=current_cost,
        recurring_annual_savings=current_cost if disposition in {"retire", "replace", "consolidate"} else 0.0,
    )
    eng["savings"]["opportunities"][opp["id"]] = opp
    if eng["phase_progress"].get("savings") == "not_started":
        eng["phase_progress"]["savings"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id) + f"#opp-{opp['id']}")


@app.route("/engagements/<engagement_id>/savings/<opp_id>/edit", methods=["POST"])
def engagement_savings_edit(engagement_id, opp_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    opp = eng["savings"]["opportunities"].get(opp_id)
    if not opp:
        abort(404)

    valid_dispositions = {k for k, _l in SAVINGS_DISPOSITION_CHOICES}
    valid_statuses = {k for k, _l in SAVINGS_STATUSES}

    title = request.form.get("title", "").strip()
    if title:
        opp["title"] = title
    disp = request.form.get("disposition", "").strip()
    if disp in valid_dispositions:
        opp["disposition"] = disp
    status = request.form.get("status", "").strip()
    if status in valid_statuses:
        opp["status"] = status
    opp["current_annual_cost"] = _safe_float(request.form.get("current_annual_cost"))
    opp["recurring_annual_savings"] = _safe_float(request.form.get("recurring_annual_savings"))
    opp["one_time_savings"] = _safe_float(request.form.get("one_time_savings"))
    opp["migration_cost"] = _safe_float(request.form.get("migration_cost"))
    opp["training_cost"] = _safe_float(request.form.get("training_cost"))
    opp["notes"] = request.form.get("notes", "").strip()
    opp["updated_at"] = datetime.utcnow().isoformat() + "Z"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id) + f"#opp-{opp_id}")


@app.route("/engagements/<engagement_id>/savings/<opp_id>/status", methods=["POST"])
def engagement_savings_status(engagement_id, opp_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    opp = eng["savings"]["opportunities"].get(opp_id)
    if not opp:
        abort(404)
    status = request.form.get("status", "").strip()
    valid_statuses = {k for k, _l in SAVINGS_STATUSES}
    if status in valid_statuses:
        opp["status"] = status
        opp["updated_at"] = datetime.utcnow().isoformat() + "Z"
        storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id) + f"#opp-{opp_id}")


@app.route("/engagements/<engagement_id>/savings/<opp_id>/delete", methods=["POST"])
def engagement_savings_delete(engagement_id, opp_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    eng["savings"]["opportunities"].pop(opp_id, None)
    storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/savings/finalize", methods=["POST"])
def engagement_savings_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["savings"]["finalized"] = False
        eng["savings"]["finalized_at"] = None
        eng["phase_progress"]["savings"] = "in_progress"
        eng["status"] = "savings"
    else:
        eng["savings"]["finalized"] = True
        eng["savings"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["savings"] = "complete"
        eng["status"] = "validation"
        if eng["phase_progress"].get("validation") == "not_started":
            eng["phase_progress"]["validation"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_savings", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/savings/estimate")
def engagement_savings_estimate(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    opps = sorted(
        eng["savings"]["opportunities"].values(),
        key=lambda o: (
            {"approved": 0, "proposed": 1, "rejected": 2}.get(o.get("status"), 3),
            -float(o.get("recurring_annual_savings") or 0),
        ),
    )
    annotated = []
    for o in opps:
        details = [
            {"product_name": products_by_id.get(pid, {}).get("product_name", ""),
             "vendor": products_by_id.get(pid, {}).get("vendor", "")}
            for pid in o.get("product_ids", []) if pid in products_by_id
        ]
        annotated.append({**o, "totals": _opportunity_totals(o), "products": details})
    summary = _savings_summary(eng)
    return render_template(
        "savings_estimate.html",
        eng=eng, opportunities=annotated, summary=summary,
        disposition_label={k: l for k, l in SAVINGS_DISPOSITION_CHOICES},
        status_label={k: l for k, l in SAVINGS_STATUSES},
    )


@app.route("/engagements/<engagement_id>/savings/estimate.csv")
def engagement_savings_estimate_csv(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_savings(eng)
    products_by_id = {p["id"]: p for p in eng.get("inventory", {}).get("products", [])}
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([
        "Title", "Status", "Disposition", "Products", "Source",
        "Current annual cost", "Recurring annual savings",
        "One-time savings", "Migration cost", "Training cost",
        "Net first-year savings", "Notes",
    ])
    disposition_label = {k: l for k, l in SAVINGS_DISPOSITION_CHOICES}
    status_label = {k: l for k, l in SAVINGS_STATUSES}
    for o in sorted(
        eng["savings"]["opportunities"].values(),
        key=lambda x: ({"approved": 0, "proposed": 1, "rejected": 2}.get(x.get("status"), 3),
                       -float(x.get("recurring_annual_savings") or 0)),
    ):
        names = [products_by_id.get(pid, {}).get("product_name", "")
                 for pid in o.get("product_ids", []) if pid in products_by_id]
        t = _opportunity_totals(o)
        writer.writerow([
            o.get("title", ""),
            status_label.get(o.get("status", ""), o.get("status", "")),
            disposition_label.get(o.get("disposition", ""), o.get("disposition", "")),
            "; ".join(names),
            o.get("source", ""),
            f"{float(o.get('current_annual_cost') or 0):.2f}",
            f"{float(o.get('recurring_annual_savings') or 0):.2f}",
            f"{float(o.get('one_time_savings') or 0):.2f}",
            f"{float(o.get('migration_cost') or 0):.2f}",
            f"{float(o.get('training_cost') or 0):.2f}",
            f"{t['first_year']:.2f}",
            o.get("notes", ""),
        ])
    fn = f"cost-savings-estimate-{eng.get('id')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 7 — Identify Technical Debt
# ---------------------------------------------------------------------------

def _ensure_tech_debt(eng):
    if "tech_debt" in eng and "flags" in eng["tech_debt"]:
        return False
    eng["tech_debt"] = {
        "flags": {},  # product_id -> {flags: [...], severity: "", notes: "", updated_at: "..."}
        "finalized": False,
        "finalized_at": None,
    }
    return True


def _suggest_debt_flags(eng):
    """
    Auto-suggest which flags might apply to each product, based on data already
    captured. Suggestions are NOT applied automatically — the user decides.
    Returns {product_id: [flag_keys]}.
    """
    products = eng.get("inventory", {}).get("products", [])
    overlap_categories = set()
    cat_counts = {}
    for p in products:
        cat = (p.get("category") or "").strip()
        if cat:
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
    overlap_categories = {c for c, n in cat_counts.items() if n >= 2}

    annual_costs = []
    for p in products:
        ac = _safe_float(p.get("total_annual_cost"))
        if ac > 0:
            annual_costs.append(ac)
    high_cost_threshold = 0
    if annual_costs:
        sorted_costs = sorted(annual_costs)
        # 75th percentile (top quartile) as the "high cost" threshold
        idx = int(len(sorted_costs) * 0.75)
        high_cost_threshold = sorted_costs[min(idx, len(sorted_costs) - 1)]

    today_iso = datetime.utcnow().date().isoformat()

    suggestions = {}
    for p in products:
        suggested = []
        # no_owner
        if not (p.get("business_owner") and p.get("technical_owner") and p.get("contract_owner")):
            suggested.append("no_owner")
        # unclear_mission
        if not (p.get("primary_use_case") or "").strip():
            suggested.append("unclear_mission")
        # duplicative
        cat = (p.get("category") or "").strip()
        if cat in overlap_categories:
            suggested.append("duplicative")
        # unused / poor_adoption  — separate signals
        purchased = _safe_int(p.get("licenses_purchased"))
        assigned = _safe_int(p.get("licenses_assigned"))
        users = _safe_int(p.get("active_users"))
        if assigned > 0 and users == 0:
            suggested.append("unused")
        elif assigned > 0 and users > 0 and users < assigned * 0.2:
            suggested.append("poor_adoption")
        # outside_governance — heuristic on purchase_source
        src = (p.get("purchase_source") or "").lower()
        if any(tag in src for tag in ("card", "personal", "shadow", "expense")):
            suggested.append("outside_governance")
        # high_cost_low_value
        ac = _safe_float(p.get("total_annual_cost"))
        if (high_cost_threshold > 0
                and ac >= high_cost_threshold
                and purchased > 0
                and users < purchased * 0.5):
            suggested.append("high_cost_low_value")
        # renewal_no_justification — within 90 days AND missing primary_use_case
        rd = (p.get("renewal_date") or "").strip()
        if rd and rd >= today_iso:
            try:
                from datetime import datetime as _dt
                rdate = _dt.strptime(rd, "%Y-%m-%d").date()
                today = _dt.utcnow().date()
                days = (rdate - today).days
                if 0 <= days <= 90 and not (p.get("primary_use_case") or "").strip():
                    suggested.append("renewal_no_justification")
            except ValueError:
                pass
        suggestions[p["id"]] = suggested
    return suggestions


def _tech_debt_summary(eng):
    flags_map = eng.get("tech_debt", {}).get("flags", {})
    products = eng.get("inventory", {}).get("products", [])
    pids_with_debt = [pid for pid, rec in flags_map.items() if rec.get("flags")]
    flag_counts = {key: 0 for key, _label, _help in TECH_DEBT_FLAGS}
    severity_counts = {"low": 0, "medium": 0, "high": 0}
    debt_annual = 0.0
    products_by_id = {p["id"]: p for p in products}
    for pid in pids_with_debt:
        rec = flags_map.get(pid, {})
        for f in rec.get("flags", []):
            if f in flag_counts:
                flag_counts[f] += 1
        sev = (rec.get("severity") or "").lower()
        if sev in severity_counts:
            severity_counts[sev] += 1
        p = products_by_id.get(pid)
        if p:
            debt_annual += _safe_float(p.get("total_annual_cost"))
    return {
        "products_with_debt": len(pids_with_debt),
        "total_products": len(products),
        "flag_counts": flag_counts,
        "severity_counts": severity_counts,
        "debt_annual": debt_annual,
    }


@app.route("/engagements/<engagement_id>/tech-debt", methods=["GET"])
def engagement_tech_debt(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if _ensure_tech_debt(eng):
        if (eng["phase_progress"].get("ai_review") == "complete"
                and eng["phase_progress"].get("tech_debt") == "not_started"):
            eng["phase_progress"]["tech_debt"] = "in_progress"
        storage.save_engagement(eng)

    products = eng.get("inventory", {}).get("products", [])
    suggestions = _suggest_debt_flags(eng)
    summary = _tech_debt_summary(eng)

    severity_filter = request.args.get("severity", "").strip()
    flag_filter = request.args.get("flag", "").strip()
    only_with_debt = request.args.get("only_with_debt") == "1"

    # Build a list of rows (one per product) annotated with current flags + suggestions.
    flags_map = eng["tech_debt"]["flags"]
    rows = []
    for p in products:
        rec = flags_map.get(p["id"], {})
        cur_flags = set(rec.get("flags", []))
        sug = [s for s in suggestions.get(p["id"], []) if s not in cur_flags]
        if only_with_debt and not cur_flags:
            continue
        if severity_filter and rec.get("severity") != severity_filter:
            continue
        if flag_filter and flag_filter not in cur_flags:
            continue
        rows.append({
            "product": p,
            "flags": list(cur_flags),
            "severity": rec.get("severity", ""),
            "notes": rec.get("notes", ""),
            "suggestions": sug,
        })

    return render_template(
        "tech_debt.html",
        eng=eng,
        rows=rows,
        debt_flags=TECH_DEBT_FLAGS,
        severities=DEBT_SEVERITIES,
        summary=summary,
        severity_filter=severity_filter,
        flag_filter=flag_filter,
        only_with_debt=only_with_debt,
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/tech-debt/<product_id>", methods=["POST"])
def engagement_tech_debt_save(engagement_id, product_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_tech_debt(eng)
    products = eng.get("inventory", {}).get("products", [])
    if not any(p["id"] == product_id for p in products):
        abort(404)

    selected = request.form.getlist("flags")
    valid_keys = {key for key, _l, _h in TECH_DEBT_FLAGS}
    selected = [k for k in selected if k in valid_keys]
    severity = request.form.get("severity", "").strip()
    if severity and severity not in {k for k, _l in DEBT_SEVERITIES}:
        severity = ""
    notes = request.form.get("notes", "").strip()

    if not (selected or severity or notes):
        eng["tech_debt"]["flags"].pop(product_id, None)
    else:
        eng["tech_debt"]["flags"][product_id] = {
            "flags": selected,
            "severity": severity,
            "notes": notes,
            "updated_at": datetime.utcnow().isoformat() + "Z",
        }
    if eng["phase_progress"].get("tech_debt") == "not_started":
        eng["phase_progress"]["tech_debt"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_tech_debt", engagement_id=engagement_id) + f"#row-{product_id}")


@app.route("/engagements/<engagement_id>/tech-debt/finalize", methods=["POST"])
def engagement_tech_debt_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_tech_debt(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["tech_debt"]["finalized"] = False
        eng["tech_debt"]["finalized_at"] = None
        eng["phase_progress"]["tech_debt"] = "in_progress"
        eng["status"] = "tech_debt"
    else:
        eng["tech_debt"]["finalized"] = True
        eng["tech_debt"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["tech_debt"] = "complete"
        eng["status"] = "savings"
        if eng["phase_progress"].get("savings") == "not_started":
            eng["phase_progress"]["savings"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_tech_debt", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/tech-debt/register")
def engagement_tech_debt_register(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_tech_debt(eng)
    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    flags_map = eng["tech_debt"]["flags"]
    flag_label = {k: l for k, l, _h in TECH_DEBT_FLAGS}
    severity_label = dict(DEBT_SEVERITIES)
    rows = []
    for pid, rec in flags_map.items():
        if not rec.get("flags"):
            continue
        p = products_by_id.get(pid)
        if not p:
            continue
        rows.append({
            "product": p,
            "flags": rec.get("flags", []),
            "severity": rec.get("severity", ""),
            "notes": rec.get("notes", ""),
        })
    rows.sort(key=lambda r: ({"high": 0, "medium": 1, "low": 2}.get(r["severity"], 3),
                              r["product"].get("product_name", "")))
    summary = _tech_debt_summary(eng)
    return render_template(
        "tech_debt_register.html",
        eng=eng, rows=rows, summary=summary,
        flag_label=flag_label, severity_label=severity_label,
        debt_flags=TECH_DEBT_FLAGS,
    )


@app.route("/engagements/<engagement_id>/tech-debt/register.csv")
def engagement_tech_debt_register_csv(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    _ensure_tech_debt(eng)
    products = eng.get("inventory", {}).get("products", [])
    products_by_id = {p["id"]: p for p in products}
    flag_label = {k: l for k, l, _h in TECH_DEBT_FLAGS}
    severity_label = dict(DEBT_SEVERITIES)
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow([
        "Product", "Vendor", "Category", "Annual cost",
        "Severity", "Flags", "Notes",
    ])
    for pid, rec in eng["tech_debt"]["flags"].items():
        if not rec.get("flags"):
            continue
        p = products_by_id.get(pid)
        if not p:
            continue
        writer.writerow([
            p.get("product_name", ""),
            p.get("vendor", ""),
            p.get("category", ""),
            p.get("total_annual_cost", ""),
            severity_label.get(rec.get("severity", ""), ""),
            "; ".join(flag_label.get(f, f) for f in rec.get("flags", [])),
            rec.get("notes", ""),
        ])
    fn = f"tech-debt-register-{eng.get('id')}.csv"
    return Response(
        sio.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ---------------------------------------------------------------------------
# Phase 6 — AI Assisted Comparison
# ---------------------------------------------------------------------------

@app.route("/engagements/<engagement_id>/ai-review")
def engagement_ai_review(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    if ai_service.ensure_ai_review(eng):
        if (eng["phase_progress"].get("overlap") == "complete"
                and eng["phase_progress"].get("ai_review") == "not_started"):
            eng["phase_progress"]["ai_review"] = "in_progress"
        storage.save_engagement(eng)

    has_inventory = bool(eng.get("inventory", {}).get("products"))
    has_key = ai_service.has_api_key()
    auto_run_triggered = False

    # Auto-run if: we have an API key, the inventory has products, the
    # findings are stale, and a run isn't already in flight.
    if (has_key
            and has_inventory
            and not eng["ai_review"].get("running")
            and ai_service.is_stale(eng)):
        if ai_service.kick_off_review(engagement_id):
            eng = storage.load_engagement(engagement_id)
            auto_run_triggered = True

    return render_template(
        "ai_review.html",
        eng=eng,
        has_inventory=has_inventory,
        has_key=has_key,
        is_stale=ai_service.is_stale(eng) if has_inventory else False,
        auto_run_triggered=auto_run_triggered,
        safe_fields=list(ai_service.AI_SAFE_FIELDS),
        redacted_fields=list(ai_service.AI_REDACTED_FIELDS),
        phases=PHASES,
    )


@app.route("/engagements/<engagement_id>/ai-review/run", methods=["POST"])
def engagement_ai_review_run(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    ai_service.ensure_ai_review(eng)
    if not ai_service.has_api_key():
        eng["ai_review"]["error"] = (
            "ANTHROPIC_API_KEY is not set. Export it in your environment "
            "before running the AI review."
        )
        storage.save_engagement(eng)
        return redirect(url_for("engagement_ai_review", engagement_id=engagement_id))
    if not eng.get("inventory", {}).get("products"):
        eng["ai_review"]["error"] = "No products in the inventory yet."
        storage.save_engagement(eng)
        return redirect(url_for("engagement_ai_review", engagement_id=engagement_id))
    ai_service.kick_off_review(engagement_id)
    return redirect(url_for("engagement_ai_review", engagement_id=engagement_id))


@app.route("/engagements/<engagement_id>/ai-review/finalize", methods=["POST"])
def engagement_ai_review_finalize(engagement_id):
    eng = storage.load_engagement(engagement_id)
    if not eng:
        abort(404)
    ai_service.ensure_ai_review(eng)
    action = request.form.get("action", "finalize")
    if action == "reopen":
        eng["ai_review"]["finalized"] = False
        eng["ai_review"]["finalized_at"] = None
        eng["phase_progress"]["ai_review"] = "in_progress"
        eng["status"] = "ai_review"
    else:
        eng["ai_review"]["finalized"] = True
        eng["ai_review"]["finalized_at"] = datetime.utcnow().isoformat() + "Z"
        eng["phase_progress"]["ai_review"] = "complete"
        eng["status"] = "tech_debt"
        if eng["phase_progress"].get("tech_debt") == "not_started":
            eng["phase_progress"]["tech_debt"] = "in_progress"
    storage.save_engagement(eng)
    return redirect(url_for("engagement_ai_review", engagement_id=engagement_id))


@app.template_filter("money")
def format_money(v):
    if v in (None, ""):
        return ""
    try:
        return f"${float(v):,.0f}"
    except (TypeError, ValueError):
        return str(v)


@app.template_filter("phase_label")
def phase_label_filter(key):
    return dict(PHASES).get(key, key)


@app.template_filter("phase_status_class")
def phase_status_class(value):
    return {
        "complete": "status-complete",
        "in_progress": "status-progress",
        "not_started": "status-pending",
    }.get(value, "status-pending")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5055, debug=True)
